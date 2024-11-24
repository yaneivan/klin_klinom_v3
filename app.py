from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, jsonify
from flask import send_from_directory, send_file
from flask_rangerequest import RangeRequest
import sqlite3
import os
import requests
import time
from flask_cors import CORS
import uuid
from apscheduler.schedulers.background import BackgroundScheduler
from mutagen.mp3 import MP3  # Import mutagen to get audio length
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tempfile
import os
from requests.exceptions import ConnectionError
from functools import wraps
import json
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash



app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
app.secret_key = 'your_secret_key'  # Change this to a secure secret key
app.config['UPLOAD_FOLDER'] = 'uploads'  # Folder to store uploaded MP3 files

# Transcription API base URL
TRANSCRIPTION_API_BASE_URL = 'http://localhost:4200'

# Database connection
def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

# Initialize the database
def init_db():
    conn = get_db_connection()
    with conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                password TEXT NOT NULL
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                mp3_path TEXT NOT NULL,
                transcription_id TEXT,
                transcription_status TEXT,
                transcription_result TEXT,
                created_at REAL NOT NULL,
                estimated_completion_time REAL,
                marked_for_training TEXT,  -- New column for marked chunks
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS transcription_speed (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                audio_length REAL NOT NULL,
                transcription_time REAL NOT NULL,
                FOREIGN KEY (project_id) REFERENCES projects (id)
            )
        ''')
    conn.close()

def update_transcription_status(project, conn):
    transcription_id = project['transcription_id']
    if transcription_id and project['transcription_status'] != 'completed':
        status_response = requests.get(f'{TRANSCRIPTION_API_BASE_URL}/status/{transcription_id}')
        if status_response.status_code == 200:
            status = status_response.json()['status']
            conn.execute('UPDATE projects SET transcription_status = ? WHERE id = ?', (status, project['id']))
            conn.commit()

            if status == 'in_queue':
                audio_length = get_audio_length(project['mp3_path'])
                estimated_time = estimate_completion_time(audio_length)
                conn.execute('UPDATE projects SET estimated_completion_time = ? WHERE id = ?', (estimated_time, project['id']))
                conn.commit()

            if status == 'completed':
                result_response = requests.get(f'{TRANSCRIPTION_API_BASE_URL}/transcribe/{transcription_id}')
                if result_response.status_code == 200:
                    result = result_response.json()
                    conn.execute('UPDATE projects SET transcription_result = ? WHERE id = ?', (str(result), project['id']))
                    conn.commit()

                    created_at = project['created_at']
                    transcription_time = time.time() - created_at
                    audio_length = get_audio_length(project['mp3_path'])
                    conn.execute('INSERT INTO transcription_speed (project_id, audio_length, transcription_time) VALUES (?, ?, ?)', 
                                 (project['id'], audio_length, transcription_time))
                    conn.commit()
        elif status_response.status_code == 404:
            conn.execute('UPDATE projects SET transcription_status = ? WHERE id = ?', ('failed', project['id']))
            conn.commit()

def estimate_completion_time(audio_length):
    conn = get_db_connection()
    average_speed = conn.execute('SELECT AVG(transcription_time / audio_length) FROM transcription_speed').fetchone()[0]
    conn.close()

    if average_speed:
        estimated_time = time.time() + (audio_length * average_speed)
    else:
        estimated_time = time.time() + (audio_length * 2.5)
    
    return estimated_time

def get_audio_length(mp3_path):
    audio = MP3(mp3_path)
    return audio.info.length

def send_for_transcription(mp3_path, transcription_id):
    try:
        with open(mp3_path, 'rb') as f:
            files = {'audio_file': f}
            data = {'id': transcription_id}
            response = requests.post(f'{TRANSCRIPTION_API_BASE_URL}/transcribe', files=files, data=data)
            if response.status_code == 200:
                status = response.json()['status']
                audio_length = get_audio_length(mp3_path)
                conn = get_db_connection()
                average_speed = conn.execute('SELECT AVG(transcription_time / audio_length) AS avg_speed FROM transcription_speed').fetchone()['avg_speed']
                conn.close()

                if average_speed:
                    estimated_time_seconds = time.time() + (audio_length * average_speed)
                else:
                    estimated_time_seconds = time.time() + audio_length

                estimated_time = time.strftime('%H:%M', time.localtime(estimated_time_seconds))
                return status, estimated_time
            elif response.status_code == 404:
                return 'failed', None
            else:
                return None, None
    except ConnectionError as e:
        print(f"Connection error: {e}")
        return 'failed', None

def check_transcription_status():
    conn = get_db_connection()
    projects = conn.execute('SELECT * FROM projects WHERE transcription_status != ?', ('completed',)).fetchall()
    for project in projects:
        update_transcription_status(project, conn)
    conn.close()

scheduler = BackgroundScheduler()
scheduler.add_job(check_transcription_status, 'interval', seconds=30)  # Adjust the interval as needed
scheduler.start()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))

        # Verify if the user exists in the database
        conn = get_db_connection()
        user = conn.execute('SELECT id FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        conn.close()

        if not user:
            # User doesn't exist, clear the session and redirect to login
            session.clear()
            return redirect(url_for('login'))

        return f(*args, **kwargs)
    return decorated_function


@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('projects'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            return redirect(url_for('projects'))
        else:
            return render_template('login.html', error='Invalid username or password')

    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if not username or not password:
            return render_template('register.html', error='Username and password are required')

        conn = get_db_connection()
        existing_user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if existing_user:
            conn.close()
            return render_template('register.html', error='Username already exists')

        hashed_password = generate_password_hash(password)
        conn.execute('INSERT INTO users (username, password) VALUES (?, ?)', (username, hashed_password))
        conn.commit()
        conn.close()

        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/projects', methods=['GET', 'POST'])
@login_required
def projects():
    user_id = session['user_id']
    conn = get_db_connection()

    user = conn.execute('SELECT username FROM users WHERE id = ?', (user_id,)).fetchone()
    username = user['username'] if user else 'Unknown'

    if request.method == 'POST':
        if 'name' in request.form and 'mp3' in request.files:
            name = request.form['name']
            mp3 = request.files['mp3']
            if mp3:
                filename = os.path.join(app.config['UPLOAD_FOLDER'], mp3.filename)
                mp3.save(filename)
                transcription_id = str(uuid.uuid4())
                status, estimated_time = send_for_transcription(filename, transcription_id)
                if status == 'in_queue' or status == 'failed':
                    created_at = time.time()
                    conn.execute('INSERT INTO projects (user_id, name, mp3_path, transcription_id, transcription_status, created_at, estimated_completion_time) VALUES (?, ?, ?, ?, ?, ?, ?)',
                                (user_id, name, filename, transcription_id, status, created_at, estimated_time))
                    conn.commit()

        elif 'delete' in request.form:
            project_id = request.form['delete']
            conn.execute('DELETE FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id))
            conn.commit()

    projects = conn.execute('SELECT * FROM projects WHERE user_id = ?', (user_id,)).fetchall()

    # Fetch average transcription speed
    average_speed = conn.execute('SELECT AVG(transcription_time / audio_length) AS avg_speed FROM transcription_speed').fetchone()['avg_speed']
    if average_speed:
        average_speed_multiplier = 1 / average_speed  # Convert to multiplier of real-time
    else:
        average_speed_multiplier = None
    conn.close()

    return render_template('projects.html', projects=projects, username=username, average_speed_multiplier=average_speed_multiplier)

@app.route('/project/<int:project_id>')
@login_required
def view_project(project_id):
    user_id = session['user_id']
    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()
    if project is None:
        conn.close()
        abort(404)

    update_transcription_status(project, conn)
    conn.close()

    return render_template('view_project.html', project=project)

@app.route('/project/<int:project_id>/status', methods=['GET'])
@login_required
def get_project_status(project_id):
    user_id = session['user_id']
    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()
    conn.close()

    if project is None:
        return jsonify({'error': 'Project not found'}), 404

    return jsonify({'status': project['transcription_status']})

@app.route('/editor/<int:project_id>')
@login_required
def editor(project_id):
    user_id = session['user_id']
    conn = get_db_connection()

    user = conn.execute('SELECT username FROM users WHERE id = ?', (user_id,)).fetchone()
    username = user['username'] if user else 'Unknown'

    project = conn.execute('SELECT * FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()
    if project is None:
        conn.close()
        abort(404)

    update_transcription_status(project, conn)
    conn.close()

    project_dict = dict(project)

    return render_template('editor.html', project=project_dict, project_id=project_id, username=username)

@app.route('/project/<int:project_id>/transcription', methods=['GET'])
@login_required
def get_transcription(project_id):
    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
    conn.close()
    if project and project['transcription_result']:
        return jsonify(eval(project['transcription_result']))
    return jsonify({'error': 'Transcription not found'}), 404

@app.route('/uploads/<path:filename>')
def download_file(filename):
    file_path = os.path.join(filename)
    range_header = request.headers.get('Range', None)
    
    app.logger.debug(f'Range header: {range_header}')
    
    if not range_header:
        return send_file(file_path)

    try:
        with open(file_path, 'rb') as f:
            range_request = RangeRequest(f, range_header)
            return range_request.make_response()
    except ValueError as e:
        app.logger.error(f'Range request error: {str(e)}')
        return send_file(file_path)

@app.route('/project/<int:project_id>/update_transcription', methods=['POST'])
@login_required
def update_transcription(project_id):
    user_id = session['user_id']
    data = request.json
    app.logger.debug(f'Received data: {data}')

    updates = data.get('updates', [])

    if not updates:
        return jsonify({'error': 'Invalid data'}), 400

    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()
    if project is None:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404

    try:
        transcription_result = eval(project['transcription_result'])
    except Exception as e:
        conn.close()
        return jsonify({'error': 'Invalid transcription result format'}), 400

    for update in updates:
        chunk_index = update.get('chunk_index')
        updated_text = update.get('updated_text')
        updated_speaker = update.get('speaker')

        if chunk_index < 0 or chunk_index >= len(transcription_result):
            continue

        if updated_text:
            transcription_result[chunk_index]['text'] = updated_text
        if updated_speaker:
            transcription_result[chunk_index]['speaker'] = updated_speaker

    conn.execute('UPDATE projects SET transcription_result = ? WHERE id = ?', (str(transcription_result), project_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

def seconds_to_hms(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

@app.route('/project/<int:project_id>/export_excel', methods=['GET'])
@login_required
def export_excel(project_id):
    user_id = session['user_id']
    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()
    conn.close()

    if project is None:
        return jsonify({'error': 'Project not found'}), 404

    transcription_result = eval(project['transcription_result'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Transcription"

    headers = ['Speaker', 'Text', 'Start Time', 'End Time', 'Start Time (h:m:s)', 'End Time (h:m:s)']
    ws.append(headers)

    for entry in transcription_result:
        speaker = entry['speaker']
        text = entry['text']
        start_time = entry['timestamp'][0]
        end_time = entry['timestamp'][1]
        start_time_hms = seconds_to_hms(start_time)
        end_time_hms = seconds_to_hms(end_time)
        ws.append([speaker, text, start_time, end_time, start_time_hms, end_time_hms])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        temp_file_path = tmp.name
        wb.save(tmp.name)

    return send_file(temp_file_path, as_attachment=True, download_name=f"{project['name']}_transcription.xlsx")    

@app.route('/project/<int:project_id>/toggle_training', methods=['POST'])
@login_required
def toggle_training(project_id):
    user_id = session['user_id']
    data = request.json
    chunk_index = data.get('chunk_index')

    if chunk_index is None:
        return jsonify({'error': 'Invalid data'}), 400

    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()

    if project is None:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404

    marked_for_training = project['marked_for_training']
    if marked_for_training:
        marked_for_training = json.loads(marked_for_training)
    else:
        marked_for_training = []

    if chunk_index in marked_for_training:
        marked_for_training.remove(chunk_index)
    else:
        marked_for_training.append(chunk_index)

    conn.execute('UPDATE projects SET marked_for_training = ? WHERE id = ?', (json.dumps(marked_for_training), project_id))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'marked_for_training': marked_for_training})

@app.route('/project/<int:project_id>/marked_for_training', methods=['GET'])
@login_required
def get_marked_for_training(project_id):
    user_id = session['user_id']
    conn = get_db_connection()
    project = conn.execute('SELECT marked_for_training FROM projects WHERE id = ? AND user_id = ?', (project_id, user_id)).fetchone()
    conn.close()

    if project is None:
        return jsonify({'error': 'Project not found'}), 404

    marked_for_training = project['marked_for_training']
    if marked_for_training:
        marked_for_training = json.loads(marked_for_training)
    else:
        marked_for_training = []

    return jsonify({'marked_for_training': marked_for_training})

if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    init_db()
    app.run(host="0.0.0.0", debug=True)